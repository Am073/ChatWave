"""Ingestion service: parse → chunk → embed → Qdrant upsert.

Decisions:
- Docling is the primary parser (Decision #10). Fallback parsers for
  PDF/DOCX/XLSX/image when Docling isn't available or extraction is empty.
- Chunking is a stdlib-only split on whitespace/_fallback_chunk_text (Decision #6).
- Embeddings go through LiteLLM (Decision #8) for provider portability.
- Vectors land in a per-tenant Qdrant collection (Decision #5).
"""
from __future__ import annotations

import hashlib
import io
import re
import uuid

from qdrant_client.models import Distance, PayloadSchemaType, PointStruct, VectorParams

from app.core.config import get_settings
from app.core.db import get_qdrant_client
from app.core.logging import get_logger
from app.services.upload_service import collection_name

log = get_logger(__name__)
_settings = get_settings()


# -------- Parsing --------


def _fallback_chunk_text(text: str, max_chars: int = 1200, overlap: int = 200) -> list[dict]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    chunks: list[dict] = []
    start = 0
    idx = 0
    while start < len(text):
        end = min(start + max_chars, len(text))
        piece = text[start:end]
        chunks.append(
            {
                "text": piece,
                "metadata": {
                    "section": f"chunk_{idx}",
                    "page": None,
                },
            }
        )
        if end == len(text):
            break
        start = max(end - overlap, start + 1)
        idx += 1
    return chunks


async def parse_to_chunks(
    file_bytes: bytes, mime_type: str, filename: str
) -> list[dict]:
    """Parse a file into chunks with metadata.

    Tries Docling first; falls back to lighter parsers. Never raises: empty
    extraction surfaces as an empty list (the task will mark the document
    failed with a clear error).
    """
    try:
        from docling.document_converter import DocumentConverter  # type: ignore[import-not-found]

        converter = DocumentConverter()
        result = converter.convert(io.BytesIO(file_bytes))
        text = result.document.export_to_text() if result.document else ""
        if text.strip():
            return _fallback_chunk_text(text)  # reuse uniform chunker
    except Exception as exc:  # noqa: BLE001
        log.info("docling_unavailable_or_failed", error=str(exc), filename=filename)

    # Fallbacks: best-effort text extract by mime type.
    text = _fallback_extract_text(file_bytes, mime_type, filename)
    return _fallback_chunk_text(text)


def _fallback_extract_text(file_bytes: bytes, mime_type: str, filename: str) -> str:
    name = filename.lower()
    try:
        if mime_type == "application/pdf" or name.endswith(".pdf"):
            import fitz  # PyMuPDF

            doc = fitz.open(stream=file_bytes, filetype="pdf")
            return "\n".join(page.get_text() for page in doc)
        if (
            mime_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or name.endswith(".docx")
        ):
            from docx import Document

            d = Document(io.BytesIO(file_bytes))
            return "\n".join(p.text for p in d.paragraphs)
        if (
            mime_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or name.endswith(".xlsx")
        ):
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" | ".join(str(c) for c in row if c is not None))
            return "\n".join(parts)
        if mime_type.startswith("image/") or name.endswith((".png", ".jpg", ".jpeg", ".webp")):
            try:
                import pytesseract
                from PIL import Image

                img = Image.open(io.BytesIO(file_bytes))
                return pytesseract.image_to_string(img)
            except Exception as exc:  # noqa: BLE001
                log.warning("ocr_unavailable", error=str(exc))
                return ""
    except Exception as exc:  # noqa: BLE001
        log.warning("fallback_parse_failed", error=str(exc), filename=filename)
    return ""


# -------- Embeddings --------


async def embed_texts(texts: list[str]) -> list[list[float]]:
    """Generate embeddings through LiteLLM (provider-portable).

    Batched: Gemini's BatchEmbedContentsRequest accepts at most 100 requests
    per call, and large PDFs chunk past that. 90 keeps a safety margin.
    Rate-limited batches (free-tier 429s) wait out Google's own backoff hint
    instead of failing the whole ingestion task.
    """
    import asyncio

    if not texts:
        return []
    import litellm

    BATCH = 90
    MAX_ATTEMPTS = 8
    out: list[list[float]] = []
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp = None
        for attempt in range(MAX_ATTEMPTS):
            try:
                resp = await litellm.aembedding(
                    model=_settings.embedding_model,
                    input=batch,
                )
                break
            except Exception as exc:  # noqa: BLE001
                msg = str(exc)
                is_throttled = "429" in msg or "RESOURCE_EXHAUSTED" in msg
                if not is_throttled or attempt == MAX_ATTEMPTS - 1:
                    raise
                wait = min(60, 5 * (2**attempt))
                log.warning(
                    "embedding_rate_limited",
                    batch_index=i // BATCH,
                    attempt=attempt + 1,
                    wait_s=wait,
                )
                await asyncio.sleep(wait)
        assert resp is not None  # unreachable: last attempt either returns or raises
        out.extend(d["embedding"] for d in resp["data"])
    return out


# -------- Qdrant indexing --------

# Payload fields the retrieval filter matches on. Qdrant (recent versions)
# rejects filtered searches with 400 unless these carry keyword indexes.
_FILTERED_FIELDS = ("collegeName", "department")


async def _ensure_collection(name: str) -> None:
    import contextlib

    client = get_qdrant_client()
    existing = await client.get_collections()
    if name not in {c.name for c in existing.collections}:
        await client.create_collection(
            collection_name=name,
            vectors_config=VectorParams(
                size=_settings.embedding_dim, distance=Distance.COSINE
            ),
        )
    for field in _FILTERED_FIELDS:
        # Idempotent: "already exists" errors are expected and ignored.
        with contextlib.suppress(Exception):
            await client.create_payload_index(
                collection_name=name,
                field_name=field,
                field_schema=PayloadSchemaType.KEYWORD,
            )


async def upsert_chunks(
    college_name: str,
    document_id: str,
    department: str | None,
    chunks: list[dict],
    embeddings: list[list[float]],
    filename: str,
) -> list[str]:
    coll = collection_name(college_name)
    await _ensure_collection(coll)
    client = get_qdrant_client()
    ids: list[str] = []
    points: list[PointStruct] = []
    for i, (chunk, vec) in enumerate(zip(chunks, embeddings, strict=False)):
        # Qdrant point IDs must be an unsigned int or a UUID — not raw hex.
        # Fold the deterministic SHA1 into a UUID so re-ingestion overwrites
        # the same points instead of duplicating them.
        digest = hashlib.sha1(f"{document_id}:{i}".encode()).hexdigest()
        point_id = str(uuid.UUID(hex=digest[:32]))
        payload = {
            "documentId": document_id,
            "chunkIndex": i,
            "text": chunk["text"],
            "collegeName": college_name,
            "department": department or "college_wide",
            "filename": filename,
            "section": chunk.get("metadata", {}).get("section"),
            "page": chunk.get("metadata", {}).get("page"),
            "embeddingModel": _settings.embedding_model,
        }
        points.append(PointStruct(id=point_id, vector=vec, payload=payload))
        ids.append(point_id)
    try:
        await client.upsert(collection_name=coll, points=points, wait=True)
    except Exception as exc:  # noqa: BLE001 - surface Qdrant's body in logs
        detail = getattr(exc, "content", None) or getattr(exc, "body", "") or ""
        log.error(
            "qdrant_upsert_failed",
            collection=coll,
            points=len(points),
            detail=str(detail)[:300] or type(exc).__name__,
        )
        raise
    return ids


async def delete_vectors(college_name: str, point_ids: list[str]) -> None:
    if not point_ids:
        return
    coll = collection_name(college_name)
    client = get_qdrant_client()
    await client.delete(
        collection_name=coll,
        points_selector={"ids": point_ids},  # type: ignore[arg-type]
        wait=True,
    )